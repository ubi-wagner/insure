"""
DBPR Structural Integrity Reserve Study (SIRS) Enricher

After the Surfside collapse (2021), Florida SB 4-D / HB 1 now requires condo
associations with buildings of 3+ stories to complete a Structural Integrity
Reserve Study (milestone inspection + reserve study) by Dec 31, 2025.

Data source — PRIMARY: Official DBPR SIRS reporting xlsx files uploaded to
filestore/System Data/DBPR/. The DBPR publishes two compliance lists that
can be downloaded from the SIRS reporting portal:

  1. "Structural Integrity Reserve Study (SIRS) Reporting*.xlsx"
     — list of associations that have FILED a SIRS report with DBPR
  2. "Structural Integrity Reserve Study (SIRS)*.xlsx"
     — master compliance list with project numbers and statuses

Data source — FALLBACK: Generates a portal lookup URL if no xlsx is found
or no match is made for an entity.

Insurance relevance:
- Associations that HAVE filed SIRS: stable, proactive boards. Good prospects.
- Associations that HAVE NOT filed: HIGH compliance risk. Special assessments
  incoming. These properties are actively shopping for new coverage because
  carriers are non-renewing non-compliant associations.

Requires: dbpr_bulk (needs dbpr_project_number or dbpr_condo_name for lookup).
"""

import logging
import os
import re
from datetime import datetime, timezone
from urllib.parse import quote_plus

from sqlalchemy.orm import Session

from agents.enrichers import record_enrichment, update_characteristics
from agents.enrichers.pipeline import register_enricher
from database.models import Entity

logger = logging.getLogger(__name__)

SIRS_PORTAL_URL = (
    "https://www2.myfloridalicense.com/condos-timeshares-mobile-homes/"
    "condominiums-and-cooperatives-sirs-reporting/"
)

# SIRS compliance deadline per FL statute 718.112(2)(g)
SIRS_DEADLINE = "2025-12-31"

# Base directory for finding uploaded xlsx files
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Search paths for SIRS xlsx files (in order of preference)
SIRS_XLSX_DIRS = [
    os.path.join(BASE_DIR, "filestore", "System Data", "DBPR"),
    os.path.join(BASE_DIR, "data"),
]

# Cache TTL — reload xlsx every 6 hours
CACHE_TTL = 3600 * 6

# In-memory caches
_filed_index: dict[str, dict] | None = None  # project_number -> filed record
_master_index: dict[str, dict] | None = None  # project_number -> master record
_name_index: dict[str, dict] | None = None    # normalized name -> record
_cache_time: float = 0


# ─── Normalization ───────────────────────────────────────────────────

def _normalize_name(name: str) -> str:
    """Normalize a condo name for matching."""
    s = name.upper()
    for noise in [
        "INC", "INC.", "LLC", "CORP", "CORP.", "LTD",
        "ASSOCIATION", "ASSN", "ASSOC",
        "OF FLORIDA", "OF FL",
        "A CONDOMINIUM", "A CONDO", "CONDOMINIUM", "CONDO",
        "NOT FOR PROFIT", "NOT-FOR-PROFIT", "NON-PROFIT", "NONPROFIT",
    ]:
        s = s.replace(noise, "")
    s = re.sub(r"[^A-Z0-9\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalize_project(project: str) -> str:
    """Normalize a DBPR project number (e.g. 'PR12345' -> 'PR12345')."""
    return re.sub(r"[^A-Z0-9]", "", project.upper().strip())


# ─── XLSX Loading ────────────────────────────────────────────────────

def _find_sirs_xlsx_files() -> tuple[str | None, str | None]:
    """Find SIRS xlsx files. Returns (filed_list_path, master_list_path)."""
    filed = None
    master = None

    for d in SIRS_XLSX_DIRS:
        if not os.path.isdir(d):
            continue
        for fname in os.listdir(d):
            lower = fname.lower()
            if not lower.endswith(".xlsx"):
                continue
            if "sirs" not in lower and "structural integrity" not in lower:
                continue
            full = os.path.join(d, fname)
            if "reporting" in lower and not filed:
                filed = full
            elif not master:
                master = full
        if filed or master:
            break

    return filed, master


# Headers we look for in the xlsx files (case-insensitive partial match)
PROJECT_HEADERS = [
    "project number", "project no", "project #", "projectno",
    "association project number", "dbpr project",
]
NAME_HEADERS = [
    "association name", "condominium name", "condo name",
    "association", "name",
]
COUNTY_HEADERS = ["county"]
STATUS_HEADERS = [
    "status", "filing status", "compliance status",
    "sirs status", "reporting status",
]
DATE_HEADERS = [
    "filed date", "submission date", "received date",
    "filing date", "completion date", "submitted",
]
ENGINEER_HEADERS = [
    "engineer", "engineering firm", "licensed professional",
    "inspector", "prepared by",
]
UNITS_HEADERS = ["units", "number of units", "unit count"]
ADDRESS_HEADERS = ["address", "street", "property address"]


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    """Find the index of a column matching any of the candidate header names."""
    if not headers:
        return None
    lowered = [str(h).strip().lower() if h else "" for h in headers]
    for i, h in enumerate(lowered):
        for cand in candidates:
            if cand in h:
                return i
    return None


def _load_sirs_xlsx() -> tuple[dict[str, dict], dict[str, dict], dict[str, dict]]:
    """Load SIRS xlsx files into project-keyed and name-keyed indices.

    Returns:
        (filed_by_project, master_by_project, by_normalized_name)
    """
    filed_by_project: dict[str, dict] = {}
    master_by_project: dict[str, dict] = {}
    by_name: dict[str, dict] = {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed — SIRS xlsx parsing disabled")
        return filed_by_project, master_by_project, by_name

    filed_path, master_path = _find_sirs_xlsx_files()

    for path, target, label in [
        (filed_path, filed_by_project, "filed"),
        (master_path, master_by_project, "master"),
    ]:
        if not path or not os.path.exists(path):
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)

            # Find header row — usually row 1, but sometimes there's a title row above
            headers = None
            data_start_row = 0
            for row_idx, row in enumerate(rows):
                if row and any(c for c in row):
                    # Check if this looks like a header row (has known column names)
                    row_lower = [str(c).strip().lower() if c else "" for c in row]
                    if any(any(cand in cell for cand in PROJECT_HEADERS + NAME_HEADERS)
                           for cell in row_lower):
                        headers = list(row)
                        data_start_row = row_idx + 1
                        break

            if not headers:
                logger.warning(f"SIRS xlsx {os.path.basename(path)}: no header row found")
                wb.close()
                continue

            proj_col = _find_column(headers, PROJECT_HEADERS)
            name_col = _find_column(headers, NAME_HEADERS)
            county_col = _find_column(headers, COUNTY_HEADERS)
            status_col = _find_column(headers, STATUS_HEADERS)
            date_col = _find_column(headers, DATE_HEADERS)
            engineer_col = _find_column(headers, ENGINEER_HEADERS)
            units_col = _find_column(headers, UNITS_HEADERS)
            addr_col = _find_column(headers, ADDRESS_HEADERS)

            logger.info(
                f"SIRS xlsx {os.path.basename(path)} headers: "
                f"project={proj_col}, name={name_col}, status={status_col}, "
                f"date={date_col}, engineer={engineer_col}"
            )

            count = 0
            # Re-iterate from the data start
            wb.close()
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx < data_start_row:
                    continue
                if not row or not any(c for c in row):
                    continue

                def get(col: int | None) -> str:
                    if col is None or col >= len(row):
                        return ""
                    val = row[col]
                    if val is None:
                        return ""
                    if isinstance(val, datetime):
                        return val.strftime("%Y-%m-%d")
                    return str(val).strip()

                project = _normalize_project(get(proj_col))
                name = get(name_col)
                if not project and not name:
                    continue

                record = {
                    "project_number": project,
                    "name": name,
                    "county": get(county_col),
                    "status": get(status_col),
                    "filed_date": get(date_col),
                    "engineer": get(engineer_col),
                    "units": get(units_col),
                    "address": get(addr_col),
                    "_source_file": os.path.basename(path),
                    "_list_type": label,  # 'filed' or 'master'
                }

                if project:
                    target[project] = record
                if name:
                    key = _normalize_name(name)
                    if key and key not in by_name:
                        by_name[key] = record
                count += 1

            wb.close()
            logger.info(
                f"SIRS xlsx {os.path.basename(path)}: loaded {count:,} records "
                f"({len(target):,} indexed by project, {len(by_name):,} by name)"
            )

        except Exception as e:
            logger.error(f"Failed to load SIRS xlsx {path}: {e}")

    return filed_by_project, master_by_project, by_name


def _get_indices() -> tuple[dict[str, dict], dict[str, dict], dict[str, dict]]:
    """Get the cached SIRS indices, reloading if stale."""
    global _filed_index, _master_index, _name_index, _cache_time
    now = datetime.now(timezone.utc).timestamp()

    if (
        _filed_index is not None
        and _master_index is not None
        and _name_index is not None
        and (now - _cache_time) < CACHE_TTL
    ):
        return _filed_index, _master_index, _name_index

    _filed_index, _master_index, _name_index = _load_sirs_xlsx()
    _cache_time = now
    return _filed_index, _master_index, _name_index


# ─── Lookup ─────────────────────────────────────────────────────────

def _lookup_sirs(project_number: str, association_name: str) -> dict | None:
    """Look up a SIRS record by project number first, then by name."""
    filed, master, by_name = _get_indices()

    if project_number:
        proj = _normalize_project(project_number)
        if proj in filed:
            return filed[proj]
        if proj in master:
            return master[proj]

    if association_name:
        key = _normalize_name(association_name)
        if key and key in by_name:
            return by_name[key]

    return None


def _build_sirs_lookup_url(project_number: str = "", association_name: str = "") -> str:
    """Build a direct lookup URL for the SIRS reporting portal."""
    if project_number:
        return f"{SIRS_PORTAL_URL}?project={quote_plus(project_number)}"
    if association_name:
        return f"{SIRS_PORTAL_URL}?name={quote_plus(association_name)}"
    return SIRS_PORTAL_URL


# ─── Enricher ───────────────────────────────────────────────────────

@register_enricher("dbpr_sirs", requires=["dbpr_bulk"])
def enrich_dbpr_sirs(entity: Entity, db: Session) -> bool:
    """Check DBPR SIRS reporting status for a condo association.

    Primary: lookup against uploaded SIRS xlsx files.
    Fallback: generate lookup URL and flag as needing manual verification.
    """
    chars = dict(entity.characteristics or {})

    project_number = str(chars.get("dbpr_project_number") or "")
    association_name = str(
        chars.get("dbpr_condo_name") or entity.name or ""
    )

    if not project_number and not association_name:
        return False

    lookup_url = _build_sirs_lookup_url(project_number, association_name)

    updates: dict = {
        "sirs_lookup_url": lookup_url,
        "sirs_deadline": SIRS_DEADLINE,
    }

    record = _lookup_sirs(project_number, association_name)

    if record:
        # Real data from the official DBPR xlsx
        list_type = record.get("_list_type", "")
        status = (record.get("status") or "").strip().lower()

        # Determine compliance based on list source and status
        if list_type == "filed":
            # On the filed list = SIRS has been submitted
            updates["sirs_completed"] = True
        elif status:
            # Status field indicates compliance
            completed_keywords = ("complete", "filed", "received", "submitted", "approved", "yes")
            non_compliant_keywords = ("pending", "not filed", "outstanding", "delinquent", "no")
            if any(k in status for k in completed_keywords):
                updates["sirs_completed"] = True
            elif any(k in status for k in non_compliant_keywords):
                updates["sirs_completed"] = False
            else:
                updates["sirs_completed"] = None  # Unknown status
        else:
            updates["sirs_completed"] = None

        if record.get("filed_date"):
            updates["sirs_completion_date"] = record["filed_date"]
        if record.get("engineer"):
            updates["sirs_engineer"] = record["engineer"]
        if status:
            updates["sirs_status"] = status

        # Compliance risk
        if updates.get("sirs_completed") is True:
            updates["sirs_compliance_risk"] = "LOW"
        elif updates.get("sirs_completed") is False:
            updates["sirs_compliance_risk"] = "HIGH"
        else:
            updates["sirs_compliance_risk"] = "UNKNOWN"

        updates["sirs_data_source"] = f"dbpr_xlsx_{list_type}"
        updates["sirs_needs_manual_verification"] = False

        logger.info(
            f"SIRS xlsx hit for entity {entity.id} '{association_name}' "
            f"(project={project_number}): completed={updates.get('sirs_completed')}, "
            f"source={updates['sirs_data_source']}"
        )

    else:
        # No xlsx hit. Default to assuming non-compliant pending verification.
        # This is consistent with the deadline being passed (Dec 31, 2025).
        updates["sirs_completed"] = False
        updates["sirs_compliance_risk"] = "HIGH"
        updates["sirs_data_source"] = "lookup_url_only"
        updates["sirs_needs_manual_verification"] = True

    update_characteristics(entity, updates, "dbpr_sirs")

    fields = [k for k, v in updates.items() if v is not None and k != "sirs_lookup_url"]

    detail_parts = []
    if updates.get("sirs_completed") is True:
        detail_parts.append("SIRS FILED")
        if updates.get("sirs_engineer"):
            detail_parts.append(f"engineer={updates['sirs_engineer']}")
    elif updates.get("sirs_completed") is False:
        detail_parts.append("NO SIRS ON FILE")
        detail_parts.append(f"compliance_risk={updates.get('sirs_compliance_risk', 'HIGH')}")
    if updates.get("sirs_needs_manual_verification"):
        detail_parts.append("needs manual verification")

    record_enrichment(
        entity, db,
        source_id="dbpr_sirs",
        fields_updated=fields,
        source_url=lookup_url,
        detail=", ".join(detail_parts) if detail_parts else "SIRS lookup URL generated",
    )

    return True
